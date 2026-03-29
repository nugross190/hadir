"""
routers/reports.py
-------------------
Excel export endpoints. Each returns a downloadable .xlsx file.

GET /reports/teacher-summary?from=2026-03-01&to=2026-03-31
GET /reports/student-summary?class_id=5&from=2026-03-01&to=2026-03-31
GET /reports/staff-accountability?from=2026-03-01&to=2026-03-31
GET /reports/daily-recap?date=2026-03-28
"""

import io
from datetime import date

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from database import get_db
from services.report_service import (
    teacher_attendance_summary,
    student_attendance_summary,
    staff_accountability_report,
    daily_recap,
)

router = APIRouter(prefix="/reports", tags=["reports"])


# ── Shared styling ──────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1B7A43", end_color="1B7A43", fill_type="solid")
TITLE_FONT = Font(name="Calibri", size=14, bold=True)
SUBTITLE_FONT = Font(name="Calibri", size=11, color="666666")
DATA_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header_row(ws, row_num, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def style_data_cell(cell, align="left"):
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal=align, vertical="center")


def auto_width(ws, min_width=10, max_width=35):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        lengths = []
        for cell in col:
            if cell.value:
                lengths.append(len(str(cell.value)))
        if lengths:
            width = min(max(max(lengths) + 2, min_width), max_width)
            ws.column_dimensions[col_letter].width = width


def make_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── 1. Teacher Attendance Summary ───────────────────────────────────────────

@router.get("/teacher-summary")
def export_teacher_summary(
    date_from: date = Query(..., alias="from"),
    date_to: date = Query(..., alias="to"),
    db: Session = Depends(get_db),
):
    data = teacher_attendance_summary(db, date_from, date_to)

    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Guru"

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "Rekap Kehadiran Guru — SMAN 5 Garut"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Periode: {date_from} s/d {date_to}"
    ws["A2"].font = SUBTITLE_FONT

    # Headers
    headers = ["No", "Kode", "Nama Guru", "Hadir", "Tidak Hadir",
               "Terlambat", "Sakit", "Izin", "Total"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))

    # Data
    for i, t in enumerate(data, 1):
        row = i + 4
        values = [i, t["kode"], t["nama"], t["hadir"], t["tidak_hadir"],
                  t["terlambat"], t["sakit"], t["izin"], t["total"]]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            align = "center" if col != 3 else "left"
            style_data_cell(cell, align)

    auto_width(ws)
    filename = f"rekap_guru_{date_from}_{date_to}.xlsx"
    return make_response(wb, filename)


# ── 2. Student Attendance Summary ───────────────────────────────────────────

@router.get("/student-summary")
def export_student_summary(
    class_id: int = Query(...),
    date_from: date = Query(..., alias="from"),
    date_to: date = Query(..., alias="to"),
    db: Session = Depends(get_db),
):
    data = student_attendance_summary(db, class_id, date_from, date_to)

    wb = Workbook()
    ws = wb.active
    ws.title = data["class_name"]

    # Title
    ws.merge_cells("A1:I1")
    ws["A1"] = f"Rekap Kehadiran Siswa — {data['class_name']}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Periode: {date_from} s/d {date_to}"
    ws["A2"].font = SUBTITLE_FONT

    # Headers
    headers = ["No", "NIS", "Nama Siswa", "L/P", "Hadir",
               "Tidak Hadir", "Sakit", "Izin", "Alpa", "Total"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))

    # Data
    for i, s in enumerate(data["students"], 1):
        row = i + 4
        values = [i, s["nis"], s["nama"], s["gender"], s["hadir"],
                  s["tidak_hadir"], s["sakit"], s["izin"], s["alpa"], s["total"]]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            align = "left" if col in (2, 3) else "center"
            style_data_cell(cell, align)

    auto_width(ws)
    filename = f"rekap_siswa_{data['class_name']}_{date_from}_{date_to}.xlsx"
    return make_response(wb, filename)


# ── 3. Staff Accountability Report ──────────────────────────────────────────

@router.get("/staff-accountability")
def export_staff_accountability(
    date_from: date = Query(..., alias="from"),
    date_to: date = Query(..., alias="to"),
    db: Session = Depends(get_db),
):
    data = staff_accountability_report(db, date_from, date_to)

    wb = Workbook()
    ws = wb.active
    ws.title = "Akuntabilitas Staff"

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "Laporan Akuntabilitas Staff — SMAN 5 Garut"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Periode: {date_from} s/d {date_to}"
    ws["A2"].font = SUBTITLE_FONT

    # Headers
    headers = ["No", "Nama Staff", "Mulai", "Selesai",
               "TOTP Verified", "Kode TOTP", "Sesi Direkam"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))

    # Data
    for i, s in enumerate(data, 1):
        row = i + 4
        values = [i, s["staff_name"], s["started_at"], s["completed_at"],
                  "Ya" if s["totp_verified"] else "Tidak",
                  s["totp_code"], s["sessions_recorded"]]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            style_data_cell(cell, "center" if col != 2 else "left")

    auto_width(ws)
    filename = f"akuntabilitas_staff_{date_from}_{date_to}.xlsx"
    return make_response(wb, filename)


# ── 4. Daily Recap ──────────────────────────────────────────────────────────

@router.get("/daily-recap")
def export_daily_recap(
    target_date: date = Query(..., alias="date"),
    db: Session = Depends(get_db),
):
    data = daily_recap(db, target_date)

    wb = Workbook()
    ws = wb.active
    ws.title = str(target_date)

    # Title
    ws.merge_cells("A1:L1")
    ws["A1"] = f"Rekap Harian — SMAN 5 Garut"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Tanggal: {target_date}"
    ws["A2"].font = SUBTITLE_FONT

    # Headers
    headers = ["No", "Kelas", "JP", "Guru", "Mapel", "Status Guru",
               "Catatan", "Jml Siswa", "Hadir", "Tidak Hadir", "Sakit", "Izin", "Alpa"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))

    # Data
    for i, r in enumerate(data, 1):
        row = i + 4
        values = [i, r["class_name"], r["period"], r["teacher_name"],
                  r["subject"], r["teacher_status"], r["teacher_notes"],
                  r["total_students"], r["hadir"], r["tidak_hadir"],
                  r["sakit"], r["izin"], r["alpa"]]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            align = "left" if col in (2, 4, 5, 7) else "center"
            style_data_cell(cell, align)

    auto_width(ws)
    filename = f"rekap_harian_{target_date}.xlsx"
    return make_response(wb, filename)

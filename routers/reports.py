"""
routers/reports.py
-------------------
Excel export endpoints. Each returns a downloadable .xlsx file.

GET /reports/teacher-summary?from=2026-03-01&to=2026-03-31
GET /reports/student-summary?class_id=5&from=2026-03-01&to=2026-03-31
GET /reports/staff-accountability?from=2026-03-01&to=2026-03-31
GET /reports/daily-recap?date=2026-03-28
"""

import io
from datetime import date

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from database import get_db
from services.report_service import (
    teacher_attendance_summary,
    student_attendance_summary,
    staff_accountability_report,
    daily_recap,
    pks_flag_ceremony_day,
    pks_flag_ceremony_summary,
)

router = APIRouter(prefix="/reports", tags=["reports"])


# ── Shared styling ──────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1B7A43", end_color="1B7A43", fill_type="solid")
TITLE_FONT = Font(name="Calibri", size=14, bold=True)
SUBTITLE_FONT = Font(name="Calibri", size=11, color="666666")
DATA_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header_row(ws, row_num, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def style_data_cell(cell, align="left"):
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal=align, vertical="center")


def auto_width(ws, min_width=10, max_width=35):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        lengths = []
        for cell in col:
            if cell.value:
                lengths.append(len(str(cell.value)))
        if lengths:
            width = min(max(max(lengths) + 2, min_width), max_width)
            ws.column_dimensions[col_letter].width = width


def make_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── 1. Teacher Attendance Summary ───────────────────────────────────────────

@router.get("/teacher-summary")
def export_teacher_summary(
    date_from: date = Query(..., alias="from"),
    date_to: date = Query(..., alias="to"),
    db: Session = Depends(get_db),
):
    data = teacher_attendance_summary(db, date_from, date_to)

    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Guru"

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "Rekap Kehadiran Guru — SMAN 5 Garut"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Periode: {date_from} s/d {date_to}"
    ws["A2"].font = SUBTITLE_FONT

    # Headers
    headers = ["No", "Kode", "Nama Guru", "Hadir", "Tidak Hadir",
               "Terlambat", "Sakit", "Izin", "Total"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))

    # Data
    for i, t in enumerate(data, 1):
        row = i + 4
        values = [i, t["kode"], t["nama"], t["hadir"], t["tidak_hadir"],
                  t["terlambat"], t["sakit"], t["izin"], t["total"]]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            align = "center" if col != 3 else "left"
            style_data_cell(cell, align)

    auto_width(ws)
    filename = f"rekap_guru_{date_from}_{date_to}.xlsx"
    return make_response(wb, filename)


# ── 2. Student Attendance Summary ───────────────────────────────────────────

@router.get("/student-summary")
def export_student_summary(
    class_id: int = Query(...),
    date_from: date = Query(..., alias="from"),
    date_to: date = Query(..., alias="to"),
    db: Session = Depends(get_db),
):
    data = student_attendance_summary(db, class_id, date_from, date_to)

    wb = Workbook()
    ws = wb.active
    ws.title = data["class_name"]

    # Title
    ws.merge_cells("A1:I1")
    ws["A1"] = f"Rekap Kehadiran Siswa — {data['class_name']}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Periode: {date_from} s/d {date_to}"
    ws["A2"].font = SUBTITLE_FONT

    # Headers
    headers = ["No", "NIS", "Nama Siswa", "L/P", "Hadir",
               "Tidak Hadir", "Sakit", "Izin", "Alpa", "Total"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))

    # Data
    for i, s in enumerate(data["students"], 1):
        row = i + 4
        values = [i, s["nis"], s["nama"], s["gender"], s["hadir"],
                  s["tidak_hadir"], s["sakit"], s["izin"], s["alpa"], s["total"]]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            align = "left" if col in (2, 3) else "center"
            style_data_cell(cell, align)

    auto_width(ws)
    filename = f"rekap_siswa_{data['class_name']}_{date_from}_{date_to}.xlsx"
    return make_response(wb, filename)


# ── 3. Staff Accountability Report ──────────────────────────────────────────

@router.get("/staff-accountability")
def export_staff_accountability(
    date_from: date = Query(..., alias="from"),
    date_to: date = Query(..., alias="to"),
    db: Session = Depends(get_db),
):
    data = staff_accountability_report(db, date_from, date_to)

    wb = Workbook()
    ws = wb.active
    ws.title = "Akuntabilitas Staff"

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "Laporan Akuntabilitas Staff — SMAN 5 Garut"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Periode: {date_from} s/d {date_to}"
    ws["A2"].font = SUBTITLE_FONT

    # Headers
    headers = ["No", "Nama Staff", "Mulai", "Selesai",
               "TOTP Verified", "Kode TOTP", "Sesi Direkam"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))

    # Data
    for i, s in enumerate(data, 1):
        row = i + 4
        values = [i, s["staff_name"], s["started_at"], s["completed_at"],
                  "Ya" if s["totp_verified"] else "Tidak",
                  s["totp_code"], s["sessions_recorded"]]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            style_data_cell(cell, "center" if col != 2 else "left")

    auto_width(ws)
    filename = f"akuntabilitas_staff_{date_from}_{date_to}.xlsx"
    return make_response(wb, filename)


# ── 4. Daily Recap ──────────────────────────────────────────────────────────

@router.get("/daily-recap")
def export_daily_recap(
    target_date: date = Query(..., alias="date"),
    db: Session = Depends(get_db),
):
    data = daily_recap(db, target_date)

    wb = Workbook()
    ws = wb.active
    ws.title = str(target_date)

    # Title
    ws.merge_cells("A1:L1")
    ws["A1"] = f"Rekap Harian — SMAN 5 Garut"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Tanggal: {target_date}"
    ws["A2"].font = SUBTITLE_FONT

    # Headers
    headers = ["No", "Kelas", "JP", "Guru", "Mapel", "Status Guru",
               "Catatan", "Jml Siswa", "Hadir", "Tidak Hadir", "Sakit", "Izin", "Alpa"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))

    # Data
    for i, r in enumerate(data, 1):
        row = i + 4
        values = [i, r["class_name"], r["period"], r["teacher_name"],
                  r["subject"], r["teacher_status"], r["teacher_notes"],
                  r["total_students"], r["hadir"], r["tidak_hadir"],
                  r["sakit"], r["izin"], r["alpa"]]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            align = "left" if col in (2, 4, 5, 7) else "center"
            style_data_cell(cell, align)

    auto_width(ws)
    filename = f"rekap_harian_{target_date}.xlsx"
    return make_response(wb, filename)


# ── 5. PKS Flag Ceremony — Current Day ──────────────────────────────────────

_DAY_ID = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"]
_STATUS_LABEL = {
    "hadir": "Hadir",
    "tidak_hadir": "Tidak Hadir",
    "izin": "Izin",
    "sakit": "Sakit",
    "alpa": "Alpa",
}


@router.get("/pks-flag-today")
def export_pks_flag_today(
    target_date: date = Query(None, alias="date"),
    db: Session = Depends(get_db),
):
    """
    Download today's (or a specific Monday's) PKS flag ceremony
    (Upacara Bendera) attendance — per-student detail, grouped by class.
    Intended for admin use.
    """
    if target_date is None:
        target_date = date.today()
    data = pks_flag_ceremony_day(db, target_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "Upacara Bendera"

    is_monday = target_date.weekday() == 0
    day_id = _DAY_ID[target_date.weekday()]

    ws.merge_cells("A1:G1")
    ws["A1"] = "Presensi Upacara Bendera (PKS) — SMAN 5 Garut"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        f"Tanggal: {target_date} ({day_id})"
        + ("" if is_monday else " — bukan hari Senin")
    )
    ws["A2"].font = SUBTITLE_FONT

    row = 4
    if not data["classes"]:
        ws.cell(row=row, column=1, value="Belum ada data presensi PKS untuk tanggal ini.")
        ws.cell(row=row, column=1).font = SUBTITLE_FONT
    else:
        # Grand-total summary
        gt = data["grand_totals"]
        ws.cell(row=row, column=1, value="Total Keseluruhan").font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"Hadir: {gt['hadir']}")
        ws.cell(row=row, column=3, value=f"Tidak Hadir: {gt['tidak_hadir']}")
        ws.cell(row=row, column=4, value=f"Izin: {gt['izin']}")
        ws.cell(row=row, column=5, value=f"Sakit: {gt['sakit']}")
        ws.cell(row=row, column=6, value=f"Alpa: {gt['alpa']}")
        ws.cell(row=row, column=7, value=f"Total Siswa: {gt['total']}")
        row += 2

        headers = ["No", "NIS", "Nama Siswa", "L/P", "Status", "Catatan"]

        for cls in data["classes"]:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            checked_at = ""
            if cls["checked_at"]:
                checked_at = f" · Direkam {cls['checked_at'][11:16]}"
            t = cls["totals"]
            cell = ws.cell(
                row=row, column=1,
                value=(
                    f"{cls['class_name']}  —  "
                    f"Hadir {t['hadir']} | Tidak Hadir {t['tidak_hadir']} | "
                    f"Izin {t['izin']} | Sakit {t['sakit']} | Alpa {t['alpa']} | "
                    f"Total {t['total']}{checked_at}"
                ),
            )
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill(start_color="6B46C1", end_color="6B46C1", fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            row += 1

            for col, h in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=h)
            style_header_row(ws, row, len(headers))
            row += 1

            for i, s in enumerate(cls["students"], 1):
                values = [
                    i, s["nis"], s["nama"], s["gender"],
                    _STATUS_LABEL.get(s["status"], s["status"]),
                    s["notes"] or "",
                ]
                for col, v in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=v)
                    align = "left" if col in (3, 6) else "center"
                    style_data_cell(cell, align)
                row += 1

            if cls.get("notes"):
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                note_cell = ws.cell(row=row, column=1, value=f"Catatan Patroli: {cls['notes']}")
                note_cell.font = SUBTITLE_FONT
                row += 1

            row += 1  # spacer between classes

    auto_width(ws)
    filename = f"upacara_bendera_{target_date}.xlsx"
    return make_response(wb, filename)


# ── 6. PKS Flag Ceremony — Period Summary (Mondays only) ────────────────────

@router.get("/pks-flag-monday")
def export_pks_flag_monday_summary(
    date_from: date = Query(..., alias="from"),
    date_to: date = Query(..., alias="to"),
    db: Session = Depends(get_db),
):
    """
    Aggregated PKS flag-ceremony (Upacara Bendera) attendance per class,
    counting only Monday check_dates within the given range.
    """
    data = pks_flag_ceremony_summary(db, date_from, date_to)

    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Upacara"

    ws.merge_cells("A1:J1")
    ws["A1"] = "Rekap Presensi Upacara Bendera (PKS) — SMAN 5 Garut"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        f"Periode Senin: {date_from} s/d {date_to} — "
        f"{len(data['mondays'])} hari Senin tercatat"
    )
    ws["A2"].font = SUBTITLE_FONT

    headers = [
        "No", "Kelas", "Tingkat", "Senin Tercatat",
        "Hadir", "Tidak Hadir", "Izin", "Sakit", "Alpa", "Total",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))

    if not data["classes"]:
        ws.cell(row=5, column=1, value="Tidak ada presensi PKS pada hari Senin di periode ini.")
    else:
        for i, c in enumerate(data["classes"], 1):
            grade = {10: "X", 11: "XI", 12: "XII"}.get(c["grade_level"], str(c["grade_level"]))
            row = i + 4
            values = [
                i, c["class_name"], grade, c["mondays_checked"],
                c["hadir"], c["tidak_hadir"], c["izin"], c["sakit"], c["alpa"], c["total"],
            ]
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=v)
                align = "left" if col == 2 else "center"
                style_data_cell(cell, align)

    auto_width(ws)
    filename = f"rekap_upacara_bendera_{date_from}_{date_to}.xlsx"
    return make_response(wb, filename)
